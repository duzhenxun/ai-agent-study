#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
整理桌面文件到Excel表格
"""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path

def get_desktop_files():
    """获取桌面文件信息"""
    desktop_path = Path.home() / 'Desktop'
    files_info = []
    
    for item in desktop_path.iterdir():
        # 跳过隐藏文件（以.开头）
        if item.name.startswith('.'):
            continue
            
        try:
            stat = item.stat()
            file_type = '文件夹' if item.is_dir() else '文件'
            extension = item.suffix if not item.is_dir() else ''
            
            # 获取文件大小（如果是文件夹，显示为0）
            size = stat.st_size if not item.is_dir() else 0
            
            # 格式化文件大小
            if size == 0:
                size_str = '0 B'
            elif size < 1024:
                size_str = f'{size} B'
            elif size < 1024 * 1024:
                size_str = f'{size/1024:.1f} KB'
            elif size < 1024 * 1024 * 1024:
                size_str = f'{size/(1024*1024):.1f} MB'
            else:
                size_str = f'{size/(1024*1024*1024):.1f} GB'
            
            # 获取修改时间
            mtime = datetime.fromtimestamp(stat.st_mtime)
            
            files_info.append({
                '文件名': item.name,
                '类型': file_type,
                '扩展名': extension,
                '大小': size_str,
                '大小(字节)': size,
                '修改时间': mtime.strftime('%Y-%m-%d %H:%M:%S'),
                '完整路径': str(item)
            })
        except Exception as e:
            print(f"处理文件 {item.name} 时出错: {e}")
    
    return files_info

def create_excel_file(files_info):
    """创建Excel文件"""
    if not files_info:
        print("桌面没有找到文件")
        return None
    
    # 创建DataFrame
    df = pd.DataFrame(files_info)
    
    # 按类型和文件名排序
    df = df.sort_values(['类型', '文件名'])
    
    # 重置索引
    df = df.reset_index(drop=True)
    
    # 添加序号列
    df.insert(0, '序号', range(1, len(df) + 1))
    
    # 保存到Excel
    output_file = '桌面文件整理.xlsx'
    
    # 使用ExcelWriter设置格式
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='桌面文件', index=False)
        
        # 获取工作簿和工作表对象
        workbook = writer.book
        worksheet = writer.sheets['桌面文件']
        
        # 设置列宽
        column_widths = {
            'A': 8,    # 序号
            'B': 40,   # 文件名
            'C': 10,   # 类型
            'D': 10,   # 扩展名
            'E': 15,   # 大小
            'F': 15,   # 修改时间
            'G': 60,   # 完整路径
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置标题行格式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行格式
        data_alignment = Alignment(vertical='center')
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.alignment = data_alignment
        
        # 添加统计信息
        total_files = len(df[df['类型'] == '文件'])
        total_folders = len(df[df['类型'] == '文件夹'])
        total_size = df['大小(字节)'].sum()
        
        # 格式化总大小
        if total_size < 1024:
            total_size_str = f'{total_size} B'
        elif total_size < 1024 * 1024:
            total_size_str = f'{total_size/1024:.1f} KB'
        elif total_size < 1024 * 1024 * 1024:
            total_size_str = f'{total_size/(1024*1024):.1f} MB'
        else:
            total_size_str = f'{total_size/(1024*1024*1024):.1f} GB'
        
        # 添加统计信息到新工作表
        stats_df = pd.DataFrame({
            '统计项': ['文件总数', '文件夹总数', '总大小', '生成时间'],
            '数值': [total_files, total_folders, total_size_str, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        })
        
        stats_df.to_excel(writer, sheet_name='统计信息', index=False)
        
        # 设置统计信息工作表格式
        stats_ws = writer.sheets['统计信息']
        stats_ws.column_dimensions['A'].width = 20
        stats_ws.column_dimensions['B'].width = 30
        
        for cell in stats_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    print(f"Excel文件已创建: {output_file}")
    print(f"共整理了 {len(df)} 个项目（{total_files} 个文件，{total_folders} 个文件夹）")
    print(f"总大小: {total_size_str}")
    
    return output_file

def main():
    """主函数"""
    print("正在扫描桌面文件...")
    files_info = get_desktop_files()
    
    if files_info:
        excel_file = create_excel_file(files_info)
        if excel_file:
            print(f"\n文件已保存到: {os.path.abspath(excel_file)}")
            print("\n桌面文件列表:")
            for i, file_info in enumerate(files_info, 1):
                print(f"{i:2d}. {file_info['文件名']} ({file_info['类型']})")
    else:
        print("桌面没有找到文件")

if __name__ == '__main__':
    main()